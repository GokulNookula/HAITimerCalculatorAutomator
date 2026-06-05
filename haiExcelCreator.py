"""
Create a Handshake AI earnings Excel tracker from a CSV task export.

Expected CSV columns by position:
1. Task ID
2. Status, ignored
3. Date
4. Total Time, expected as MM:SS. HH:MM:SS also works.
"""

from __future__ import annotations

import csv
import shutil
from datetime import datetime, timedelta, date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


defaultTimeCapsMin = {
    "250929-text-to-image-h2h": 10,
    "251016-vision-vlm-h2h": 20,
    "251210-conversations-to-response-h2h": 10,
    "260124-text-image-to-text-h2h": 15,
    "260126-text-to-image-compare": 10,
    "260209-omni-elo": 10,
    "260212-ud-caption-elo": 20,
    "260308-omni-r2i-elo": 15,
    "260310-live-s2s-elo": 45,
    "260403-omni-multiturn-elo": 20,
    "vs-1776345130-video-audio-caption-comparison-robinsr-v2-apr16": 20,
    "find-the-boundary-v2":None,
}

approvedTaskTypes = list(defaultTimeCapsMin.keys())


headerFill = PatternFill("solid", fgColor="1F4E78")
subheaderFill = PatternFill("solid", fgColor="D9EAF7")
whiteFont = Font(color="FFFFFF", bold=True)
boldFont = Font(bold=True)
thinBorder = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def parseTaskDate(value: str) -> date:
    text = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Could not parse date: {value!r}")


def parseTotalTimeToMinutes(value: str) -> float:
    """
    Converts Total Time to minutes.
    The Handshake export examples look like MM:SS, such as 10:19.
    HH:MM:SS also works.
    """
    text = str(value).strip()
    if not text:
        return 0.0

    parts = text.split(":")
    try:
        if len(parts) == 2:
            minutes = int(parts[0])
            seconds = int(parts[1])
            return minutes + seconds / 60
        if len(parts) == 3:
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = int(parts[2])
            return hours * 60 + minutes + seconds / 60
        return float(text)
    except ValueError as exc:
        raise ValueError(f"Could not parse total time: {value!r}") from exc


def weekStartMonday(taskDate: date) -> date:
    return taskDate - timedelta(days=taskDate.weekday())


def firstDayOfMonth(taskDate: date) -> date:
    return date(taskDate.year, taskDate.month, 1)


def latestCsvInFolder(folder: Path) -> Path:
    csvFiles = list(folder.glob("*.csv"))
    if not csvFiles:
        raise FileNotFoundError(f"No .csv files found in: {folder}")
    return max(csvFiles, key=lambda path: path.stat().st_mtime)


def resolveCsvPath(csvArg: str | None, outputFolder: str) -> Path:
    if csvArg:
        path = Path(csvArg)
        if path.is_dir():
            return latestCsvInFolder(path)
        return path
    return latestCsvInFolder(Path(outputFolder))


def readTasks(csvPath: Path) -> list[dict]:
    tasks = []
    with csvPath.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        header = next(reader, None)
        if not header:
            return tasks

        for lineNumber, row in enumerate(reader, start=2):
            if len(row) < 4 or not str(row[0]).strip():
                continue

            taskId = str(row[0]).strip()
            taskDate = parseTaskDate(row[2])
            totalTimeText = str(row[3]).strip()
            actualMinutes = parseTotalTimeToMinutes(totalTimeText)
            startOfWeek = weekStartMonday(taskDate)

            tasks.append(
                {
                    "taskId": taskId,
                    "date": taskDate,
                    "totalTime": totalTimeText,
                    "actualMinutes": round(actualMinutes, 4),
                    "weekStart": startOfWeek,
                    "weekEnd": startOfWeek + timedelta(days=6),
                    "monthStart": firstDayOfMonth(taskDate),
                }
            )
    return tasks


def addDefinedName(workbook: Workbook, name: str, attrText: str) -> None:
    definedName = DefinedName(name, attr_text=attrText)
    try:
        workbook.defined_names.add(definedName)
    except AttributeError:
        workbook.defined_names.append(definedName)


def styleTitle(ws, title: str, throughCol: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=throughCol)
    cell = ws.cell(row=1, column=1, value=title)
    cell.fill = headerFill
    cell.font = Font(color="FFFFFF", bold=True, size=14)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24


def styleHeaderRow(ws, rowNumber: int, startCol: int, endCol: int) -> None:
    for col in range(startCol, endCol + 1):
        cell = ws.cell(row=rowNumber, column=col)
        cell.fill = headerFill
        cell.font = whiteFont
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thinBorder


def applyBasicSheetFormatting(ws) -> None:
    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = thinBorder


def setColumnWidths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def createWorkbook(tasks: list[dict], projectName: str, hourlyPay: float, outputPath: Path) -> None:
    wb = Workbook()
    wsTasks = wb.active
    wsTasks.title = "Tasks"
    wsTaskTypes = wb.create_sheet("Task Types")
    wsSummary = wb.create_sheet("Weekly Summary")
    wsMonthly = wb.create_sheet("Monthly Summary")
    wsSettings = wb.create_sheet("Settings")

    # Settings sheet
    styleTitle(wsSettings, "Project Settings", 3)
    wsSettings.append(["Setting", "Value", "Notes"])
    wsSettings.append(["Project Name", projectName, "Change this if the project name changes."])
    wsSettings.append(["Hourly Pay Rate", hourlyPay, "Used to calculate earnings."])
    wsSettings.append([
        "Paid Time Rule",
        "MIN(actual minutes, task cap minutes)",
        "If Handshake pays the full cap no matter what, change Tasks column H formula from MIN(F,G) to G.",
    ])
    styleHeaderRow(wsSettings, 2, 1, 3)
    wsSettings["B4"].number_format = '$#,##0.00'
    setColumnWidths(wsSettings, {"A": 20, "B": 35, "C": 80})
    applyBasicSheetFormatting(wsSettings)

    # Task Types sheet
    styleTitle(wsTaskTypes, "Approved Task Types and Time Caps", 3)
    wsTaskTypes.append(["Task Type", "Time Cap Minutes", "Notes"])
    for taskType in approvedTaskTypes:
        wsTaskTypes.append([
            taskType,
            defaultTimeCapsMin.get(taskType, ""),
            "You can change this max paid time anytime.",
        ])

    styleHeaderRow(wsTaskTypes, 2, 1, 3)
    setColumnWidths(wsTaskTypes, {"A": 75, "B": 18, "C": 50})
    wsTaskTypes.freeze_panes = "A3"
    applyBasicSheetFormatting(wsTaskTypes)
    wsTaskTypes["A2"].comment = Comment(
        "Approved task types. You can edit names or add more manually below the existing list.",
        "ChatGPT",
    )
    wsTaskTypes["B2"].comment = Comment(
        "Max paid minutes for each approved task type. Changing this updates the Tasks sheet earnings formulas.",
        "ChatGPT",
    )

    if len(approvedTaskTypes) >= 1:
        taskTypeTableEnd = 2 + len(approvedTaskTypes)
        taskTypeTable = Table(displayName="TaskTypesTable", ref=f"A2:C{taskTypeTableEnd}")
        taskTypeTable.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        wsTaskTypes.add_table(taskTypeTable)

    # Named range used by the task type dropdown.
    addDefinedName(wb, "TaskTypeList", "'Task Types'!$A$3:$A$1000")

    # Tasks sheet
    taskHeaders = [
        "Task ID",
        "Date",
        "Total Time",
        "Type of Task",
        "Earning",
        "Actual Minutes",
        "Time Cap Minutes",
        "Paid Minutes",
        "Week Start",
        "Week End",
        "Project",
    ]
    wsTasks.append(taskHeaders)
    styleHeaderRow(wsTasks, 1, 1, len(taskHeaders))

    for task in tasks:
        rowNum = wsTasks.max_row + 1
        wsTasks.append(
            [
                task["taskId"],
                task["date"],
                task["totalTime"],
                "",
                None,
                task["actualMinutes"],
                None,
                None,
                task["weekStart"],
                task["weekEnd"],
                None,
            ]
        )
        wsTasks.cell(row=rowNum, column=5).value = (
            f'=IF(OR($H{rowNum}="",\'Settings\'!$B$4=""),"",$H{rowNum}/60*\'Settings\'!$B$4)'
        )
        timeCapLookupFormula = f'VLOOKUP($D{rowNum},\'Task Types\'!$A$3:$B$1000,2,FALSE)'

        wsTasks.cell(row=rowNum, column=7).value = (
            f'=IF($D{rowNum}="","",'
            f'IFERROR(IF({timeCapLookupFormula}=0,"",{timeCapLookupFormula}),""))'
        )

        wsTasks.cell(row=rowNum, column=8).value = (
            f'=IF(OR($F{rowNum}="",$G{rowNum}=""),"",MIN($F{rowNum},$G{rowNum}))'
        )
        wsTasks.cell(row=rowNum, column=11).value = "='Settings'!$B$3"

    lastTaskRow = max(wsTasks.max_row, 2)
    wsTasks.freeze_panes = "A2"
    setColumnWidths(
        wsTasks,
        {
            "A": 42,
            "B": 14,
            "C": 14,
            "D": 62,
            "E": 14,
            "F": 16,
            "G": 18,
            "H": 16,
            "I": 14,
            "J": 14,
            "K": 22,
        },
    )

    for row in range(2, lastTaskRow + 1):
        wsTasks.cell(row=row, column=2).number_format = "m/d/yyyy"
        wsTasks.cell(row=row, column=5).number_format = '$#,##0.00'
        wsTasks.cell(row=row, column=6).number_format = "0.00"
        wsTasks.cell(row=row, column=7).number_format = "0.00"
        wsTasks.cell(row=row, column=8).number_format = "0.00"
        wsTasks.cell(row=row, column=9).number_format = "m/d/yyyy"
        wsTasks.cell(row=row, column=10).number_format = "m/d/yyyy"

    taskDropdown = DataValidation(
        type="list",
        formula1="=TaskTypeList",
        allow_blank=True,
        showDropDown=False,
    )
    # The dropdown gives approved choices. If you need another type, add it manually in Task Types columns A:B.
    taskDropdown.showErrorMessage = False
    taskDropdown.error = "Choose an approved task type from the dropdown. To add a new one, add it manually in Task Types columns A:B."
    taskDropdown.errorTitle = "Task type"
    taskDropdown.prompt = "Pick a task type from the dropdown. To add a new type, add it manually in Task Types columns A:B."
    taskDropdown.promptTitle = "Type of Task"
    wsTasks.add_data_validation(taskDropdown)
    taskDropdown.add(f"D2:D{max(lastTaskRow + 500, 1000)}")

    if lastTaskRow >= 2:
        tasksTable = Table(displayName="TasksTable", ref=f"A1:K{lastTaskRow}")
        tasksTable.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        wsTasks.add_table(tasksTable)

    wsTasks["D1"].comment = Comment(
        "Dropdown comes from Task Types!A3:A1000. To add a new task type, add it manually in Task Types columns A:B, then select or type it here.",
        "ChatGPT",
    )
    wsTasks["E1"].comment = Comment(
        "Earning = paid minutes / 60 * hourly pay rate. Paid minutes uses the smaller of actual minutes and the task time cap.",
        "ChatGPT",
    )

    applyBasicSheetFormatting(wsTasks)
    wsTasks.conditional_formatting.add(
        f"E2:E{lastTaskRow}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="E2F0D9")),
    )

    # Weekly Summary sheet
    styleTitle(wsSummary, "Weekly Earnings Summary", 7)
    weeklyHeaders = [
        "Week Start",
        "Week End",
        "Week Label",
        "Total Tasks",
        "Actual Minutes",
        "Paid Minutes",
        "Total Earnings",
    ]
    wsSummary.append(weeklyHeaders)
    styleHeaderRow(wsSummary, 2, 1, len(weeklyHeaders))

    uniqueWeeks = sorted({task["weekStart"] for task in tasks})
    for weekStart in uniqueWeeks:
        rowNum = wsSummary.max_row + 1
        weekEnd = weekStart + timedelta(days=6)
        wsSummary.append([weekStart, weekEnd, None, None, None, None, None])
        wsSummary.cell(row=rowNum, column=3).value = f'=TEXT(A{rowNum},"m/d/yyyy")&" - "&TEXT(B{rowNum},"m/d/yyyy")'
        wsSummary.cell(row=rowNum, column=4).value = f'=COUNTIFS(Tasks!$I:$I,$A{rowNum})'
        wsSummary.cell(row=rowNum, column=5).value = f'=SUMIFS(Tasks!$F:$F,Tasks!$I:$I,$A{rowNum})'
        wsSummary.cell(row=rowNum, column=6).value = f'=SUMIFS(Tasks!$H:$H,Tasks!$I:$I,$A{rowNum})'
        wsSummary.cell(row=rowNum, column=7).value = f'=SUMIFS(Tasks!$E:$E,Tasks!$I:$I,$A{rowNum})'

    lastWeekRow = max(wsSummary.max_row, 3)
    setColumnWidths(wsSummary, {"A": 14, "B": 14, "C": 28, "D": 12, "E": 16, "F": 14, "G": 16, "I": 22})
    for row in range(3, lastWeekRow + 1):
        wsSummary.cell(row=row, column=1).number_format = "m/d/yyyy"
        wsSummary.cell(row=row, column=2).number_format = "m/d/yyyy"
        wsSummary.cell(row=row, column=5).number_format = "0.00"
        wsSummary.cell(row=row, column=6).number_format = "0.00"
        wsSummary.cell(row=row, column=7).number_format = '$#,##0.00'

    if len(uniqueWeeks) >= 1:
        weeklyTable = Table(displayName="WeeklySummaryTable", ref=f"A2:G{lastWeekRow}")
        weeklyTable.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        wsSummary.add_table(weeklyTable)

        weeklyChart = BarChart()
        weeklyChart.title = "Weekly Earnings"
        weeklyChart.y_axis.title = "Earnings"
        weeklyChart.x_axis.title = "Week"
        weeklyChart.height = 8
        weeklyChart.width = 18
        weeklyData = Reference(wsSummary, min_col=7, min_row=2, max_row=lastWeekRow)
        weeklyCategories = Reference(wsSummary, min_col=3, min_row=3, max_row=lastWeekRow)
        weeklyChart.add_data(weeklyData, titles_from_data=True)
        weeklyChart.set_categories(weeklyCategories)
        wsSummary.add_chart(weeklyChart, "I2")

    wsSummary.freeze_panes = "A3"
    applyBasicSheetFormatting(wsSummary)

    # Monthly Summary sheet
    styleTitle(wsMonthly, "Monthly Earnings Summary", 5)
    monthlyHeaders = ["Month", "Total Tasks", "Actual Minutes", "Paid Minutes", "Total Earnings"]
    wsMonthly.append(monthlyHeaders)
    styleHeaderRow(wsMonthly, 2, 1, len(monthlyHeaders))

    uniqueMonths = sorted({task["monthStart"] for task in tasks})
    for monthStart in uniqueMonths:
        rowNum = wsMonthly.max_row + 1
        wsMonthly.append([monthStart, None, None, None, None])
        wsMonthly.cell(row=rowNum, column=2).value = (
            f'=COUNTIFS(Tasks!$B:$B,">="&$A{rowNum},Tasks!$B:$B,"<"&EDATE($A{rowNum},1))'
        )
        wsMonthly.cell(row=rowNum, column=3).value = (
            f'=SUMIFS(Tasks!$F:$F,Tasks!$B:$B,">="&$A{rowNum},Tasks!$B:$B,"<"&EDATE($A{rowNum},1))'
        )
        wsMonthly.cell(row=rowNum, column=4).value = (
            f'=SUMIFS(Tasks!$H:$H,Tasks!$B:$B,">="&$A{rowNum},Tasks!$B:$B,"<"&EDATE($A{rowNum},1))'
        )
        wsMonthly.cell(row=rowNum, column=5).value = (
            f'=SUMIFS(Tasks!$E:$E,Tasks!$B:$B,">="&$A{rowNum},Tasks!$B:$B,"<"&EDATE($A{rowNum},1))'
        )

    lastMonthRow = max(wsMonthly.max_row, 3)
    setColumnWidths(wsMonthly, {"A": 16, "B": 12, "C": 16, "D": 14, "E": 16, "G": 22})
    for row in range(3, lastMonthRow + 1):
        wsMonthly.cell(row=row, column=1).number_format = "mmm yyyy"
        wsMonthly.cell(row=row, column=3).number_format = "0.00"
        wsMonthly.cell(row=row, column=4).number_format = "0.00"
        wsMonthly.cell(row=row, column=5).number_format = '$#,##0.00'

    if len(uniqueMonths) >= 1:
        monthlyTable = Table(displayName="MonthlySummaryTable", ref=f"A2:E{lastMonthRow}")
        monthlyTable.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        wsMonthly.add_table(monthlyTable)

        monthlyChart = BarChart()
        monthlyChart.title = "Monthly Earnings"
        monthlyChart.y_axis.title = "Earnings"
        monthlyChart.x_axis.title = "Month"
        monthlyChart.height = 8
        monthlyChart.width = 16
        monthlyData = Reference(wsMonthly, min_col=5, min_row=2, max_row=lastMonthRow)
        monthlyCategories = Reference(wsMonthly, min_col=1, min_row=3, max_row=lastMonthRow)
        monthlyChart.add_data(monthlyData, titles_from_data=True)
        monthlyChart.set_categories(monthlyCategories)
        wsMonthly.add_chart(monthlyChart, "G2")

    wsMonthly.freeze_panes = "A3"
    applyBasicSheetFormatting(wsMonthly)

    # Make the workbook open on the Tasks sheet.
    wb.active = wb.sheetnames.index("Tasks")

    outputPath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(outputPath)



def extractHourlyPayValue(value: str) -> float:
    """Extracts a numeric hourly pay value from strings like '$17' or '$17/hr'."""
    import re

    text = str(value).strip()
    payMatch = re.search(r"\d+(?:\.\d+)?", text)

    if not payMatch:
        raise ValueError(f"Could not extract hourly pay from: {value!r}")

    return float(payMatch.group())


def readProjectDetailsFromCsv(csvPath: Path) -> tuple[str, float]:
    """Reads project name and hourly pay from the CSV created by haiCSVCreator."""
    with csvPath.open("r", newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        firstRow = next(reader, None)

    if not firstRow:
        raise ValueError(f"CSV file is empty, so project details could not be read: {csvPath}")

    projectName = str(firstRow.get("Project Name", "Handshake AI Project")).strip() or "Handshake AI Project"
    hourlyPayText = firstRow.get("Project Pay Per Hour", "")
    hourlyPay = extractHourlyPayValue(hourlyPayText)

    return projectName, hourlyPay

def cleanPathInput(pathValue: str | Path | None) -> Path | None:
    if pathValue is None:
        return None

    pathText = str(pathValue).strip().strip('"').strip("'")

    if not pathText:
        return None

    return Path(pathText)


def resolveExistingWorkbookPath(
    existingWorkbookPath: str | Path | None,
    outputFolder: str,
    outputFileName: str,
) -> Path | None:
    cleanedPath = cleanPathInput(existingWorkbookPath)

    if cleanedPath is None:
        return None

    if cleanedPath.is_dir():
        return cleanedPath / outputFileName

    return cleanedPath

def resolveOutputPath(
    outputPath: str | Path | None,
    outputFolder: str,
    outputFileName: str,
) -> Path:
    cleanedOutputPath = cleanPathInput(outputPath)

    if cleanedOutputPath is not None:
        return cleanedOutputPath

    return Path(outputFolder) / outputFileName


def getHeaderMap(ws) -> dict[str, int]:
    headerMap = {}

    for cell in ws[1]:
        if cell.value is None:
            continue

        headerName = str(cell.value).strip()
        if headerName:
            headerMap[headerName] = cell.column

    return headerMap


def getExistingTaskIds(wsTasks, taskIdCol: int) -> set[str]:
    existingTaskIds = set()

    for rowNum in range(2, wsTasks.max_row + 1):
        taskId = wsTasks.cell(row=rowNum, column=taskIdCol).value

        if taskId is not None and str(taskId).strip():
            existingTaskIds.add(str(taskId).strip())

    return existingTaskIds


def updateTasksTableRange(wsTasks, lastTaskRow: int | None = None) -> None:
    if not wsTasks.tables:
        return

    if lastTaskRow is None:
        lastTaskRow = wsTasks.max_row

    endColLetter = get_column_letter(wsTasks.max_column)
    newRef = f"A1:{endColLetter}{lastTaskRow}"

    for table in wsTasks.tables.values():
        if table.displayName == "TasksTable":
            table.ref = newRef
            return

def parseExcelDateValue(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    if not text:
        return None

    for dateFormat in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, dateFormat).date()
        except ValueError:
            pass

    return None


def removeWorksheetIfExists(workbook, sheetName: str) -> None:
    if sheetName in workbook.sheetnames:
        workbook.remove(workbook[sheetName])


def getSummaryInsertIndex(workbook) -> int:
    if "Settings" in workbook.sheetnames:
        return workbook.sheetnames.index("Settings")

    return len(workbook.sheetnames)


def collectSummaryPeriodsFromTasks(wsTasks, headerMap: dict[str, int]) -> tuple[list[date], list[date]]:
    taskIdCol = headerMap["Task ID"]
    dateCol = headerMap["Date"]
    weekStartCol = headerMap["Week Start"]

    weekStartDates = set()
    monthStartDates = set()

    for rowNum in range(2, wsTasks.max_row + 1):
        taskIdValue = wsTasks.cell(row=rowNum, column=taskIdCol).value

        if taskIdValue is None or not str(taskIdValue).strip():
            continue

        taskDate = parseExcelDateValue(wsTasks.cell(row=rowNum, column=dateCol).value)
        weekStartDate = parseExcelDateValue(wsTasks.cell(row=rowNum, column=weekStartCol).value)

        if taskDate is not None:
            monthStartDates.add(firstDayOfMonth(taskDate))

        if weekStartDate is not None:
            weekStartDates.add(weekStartDate)
        elif taskDate is not None:
            weekStartDates.add(weekStartMonday(taskDate))

    return sorted(weekStartDates), sorted(monthStartDates)


def refreshWeeklySummarySheet(workbook, weekStartDates: list[date], headerMap: dict[str, int]) -> None:
    insertIndex = getSummaryInsertIndex(workbook)
    wsSummary = workbook.create_sheet("Weekly Summary", insertIndex)

    styleTitle(wsSummary, "Weekly Earnings Summary", 7)

    weeklyHeaders = [
        "Week Start",
        "Week End",
        "Week Label",
        "Total Tasks",
        "Actual Minutes",
        "Paid Minutes",
        "Total Earnings",
    ]

    wsSummary.append(weeklyHeaders)
    styleHeaderRow(wsSummary, 2, 1, len(weeklyHeaders))

    weekStartColLetter = get_column_letter(headerMap["Week Start"])
    actualMinutesColLetter = get_column_letter(headerMap["Actual Minutes"])
    paidMinutesColLetter = get_column_letter(headerMap["Paid Minutes"])
    earningColLetter = get_column_letter(headerMap["Earning"])

    for weekStartDate in weekStartDates:
        rowNum = wsSummary.max_row + 1
        weekEndDate = weekStartDate + timedelta(days=6)

        wsSummary.append([weekStartDate, weekEndDate, None, None, None, None, None])

        wsSummary.cell(row=rowNum, column=3).value = (
            f'=TEXT(A{rowNum},"m/d/yyyy")&" - "&TEXT(B{rowNum},"m/d/yyyy")'
        )
        wsSummary.cell(row=rowNum, column=4).value = (
            f'=COUNTIFS(Tasks!${weekStartColLetter}:${weekStartColLetter},$A{rowNum})'
        )
        wsSummary.cell(row=rowNum, column=5).value = (
            f'=SUMIFS(Tasks!${actualMinutesColLetter}:${actualMinutesColLetter},'
            f'Tasks!${weekStartColLetter}:${weekStartColLetter},$A{rowNum})'
        )
        wsSummary.cell(row=rowNum, column=6).value = (
            f'=SUMIFS(Tasks!${paidMinutesColLetter}:${paidMinutesColLetter},'
            f'Tasks!${weekStartColLetter}:${weekStartColLetter},$A{rowNum})'
        )
        wsSummary.cell(row=rowNum, column=7).value = (
            f'=SUMIFS(Tasks!${earningColLetter}:${earningColLetter},'
            f'Tasks!${weekStartColLetter}:${weekStartColLetter},$A{rowNum})'
        )

    lastWeekRow = max(wsSummary.max_row, 3)

    setColumnWidths(
        wsSummary,
        {
            "A": 14,
            "B": 14,
            "C": 28,
            "D": 12,
            "E": 16,
            "F": 14,
            "G": 16,
            "I": 22,
        },
    )

    for rowNum in range(3, lastWeekRow + 1):
        wsSummary.cell(row=rowNum, column=1).number_format = "m/d/yyyy"
        wsSummary.cell(row=rowNum, column=2).number_format = "m/d/yyyy"
        wsSummary.cell(row=rowNum, column=5).number_format = "0.00"
        wsSummary.cell(row=rowNum, column=6).number_format = "0.00"
        wsSummary.cell(row=rowNum, column=7).number_format = '$#,##0.00'

    if weekStartDates:
        weeklyTable = Table(displayName="WeeklySummaryTable", ref=f"A2:G{lastWeekRow}")
        weeklyTable.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        wsSummary.add_table(weeklyTable)

        weeklyChart = BarChart()
        weeklyChart.title = "Weekly Earnings"
        weeklyChart.y_axis.title = "Earnings"
        weeklyChart.x_axis.title = "Week"
        weeklyChart.height = 8
        weeklyChart.width = 18

        weeklyData = Reference(wsSummary, min_col=7, min_row=2, max_row=lastWeekRow)
        weeklyCategories = Reference(wsSummary, min_col=3, min_row=3, max_row=lastWeekRow)

        weeklyChart.add_data(weeklyData, titles_from_data=True)
        weeklyChart.set_categories(weeklyCategories)

        wsSummary.add_chart(weeklyChart, "I2")

    wsSummary.freeze_panes = "A3"
    applyBasicSheetFormatting(wsSummary)


def refreshMonthlySummarySheet(workbook, monthStartDates: list[date], headerMap: dict[str, int]) -> None:
    insertIndex = getSummaryInsertIndex(workbook)
    wsMonthly = workbook.create_sheet("Monthly Summary", insertIndex)

    styleTitle(wsMonthly, "Monthly Earnings Summary", 5)

    monthlyHeaders = [
        "Month",
        "Total Tasks",
        "Actual Minutes",
        "Paid Minutes",
        "Total Earnings",
    ]

    wsMonthly.append(monthlyHeaders)
    styleHeaderRow(wsMonthly, 2, 1, len(monthlyHeaders))

    dateColLetter = get_column_letter(headerMap["Date"])
    actualMinutesColLetter = get_column_letter(headerMap["Actual Minutes"])
    paidMinutesColLetter = get_column_letter(headerMap["Paid Minutes"])
    earningColLetter = get_column_letter(headerMap["Earning"])

    for monthStartDate in monthStartDates:
        rowNum = wsMonthly.max_row + 1

        wsMonthly.append([monthStartDate, None, None, None, None])

        wsMonthly.cell(row=rowNum, column=2).value = (
            f'=COUNTIFS(Tasks!${dateColLetter}:${dateColLetter},">="&$A{rowNum},'
            f'Tasks!${dateColLetter}:${dateColLetter},"<"&EDATE($A{rowNum},1))'
        )
        wsMonthly.cell(row=rowNum, column=3).value = (
            f'=SUMIFS(Tasks!${actualMinutesColLetter}:${actualMinutesColLetter},'
            f'Tasks!${dateColLetter}:${dateColLetter},">="&$A{rowNum},'
            f'Tasks!${dateColLetter}:${dateColLetter},"<"&EDATE($A{rowNum},1))'
        )
        wsMonthly.cell(row=rowNum, column=4).value = (
            f'=SUMIFS(Tasks!${paidMinutesColLetter}:${paidMinutesColLetter},'
            f'Tasks!${dateColLetter}:${dateColLetter},">="&$A{rowNum},'
            f'Tasks!${dateColLetter}:${dateColLetter},"<"&EDATE($A{rowNum},1))'
        )
        wsMonthly.cell(row=rowNum, column=5).value = (
            f'=SUMIFS(Tasks!${earningColLetter}:${earningColLetter},'
            f'Tasks!${dateColLetter}:${dateColLetter},">="&$A{rowNum},'
            f'Tasks!${dateColLetter}:${dateColLetter},"<"&EDATE($A{rowNum},1))'
        )

    lastMonthRow = max(wsMonthly.max_row, 3)

    setColumnWidths(
        wsMonthly,
        {
            "A": 16,
            "B": 12,
            "C": 16,
            "D": 14,
            "E": 16,
            "G": 22,
        },
    )

    for rowNum in range(3, lastMonthRow + 1):
        wsMonthly.cell(row=rowNum, column=1).number_format = "mmm yyyy"
        wsMonthly.cell(row=rowNum, column=3).number_format = "0.00"
        wsMonthly.cell(row=rowNum, column=4).number_format = "0.00"
        wsMonthly.cell(row=rowNum, column=5).number_format = '$#,##0.00'

    if monthStartDates:
        monthlyTable = Table(displayName="MonthlySummaryTable", ref=f"A2:E{lastMonthRow}")
        monthlyTable.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        wsMonthly.add_table(monthlyTable)

        monthlyChart = BarChart()
        monthlyChart.title = "Monthly Earnings"
        monthlyChart.y_axis.title = "Earnings"
        monthlyChart.x_axis.title = "Month"
        monthlyChart.height = 8
        monthlyChart.width = 16

        monthlyData = Reference(wsMonthly, min_col=5, min_row=2, max_row=lastMonthRow)
        monthlyCategories = Reference(wsMonthly, min_col=1, min_row=3, max_row=lastMonthRow)

        monthlyChart.add_data(monthlyData, titles_from_data=True)
        monthlyChart.set_categories(monthlyCategories)

        wsMonthly.add_chart(monthlyChart, "G2")

    wsMonthly.freeze_panes = "A3"
    applyBasicSheetFormatting(wsMonthly)


def refreshSummarySheets(workbook) -> None:
    if "Tasks" not in workbook.sheetnames:
        return

    wsTasks = workbook["Tasks"]
    headerMap = getHeaderMap(wsTasks)

    requiredHeaders = [
        "Task ID",
        "Date",
        "Earning",
        "Actual Minutes",
        "Paid Minutes",
        "Week Start",
    ]

    missingHeaders = [header for header in requiredHeaders if header not in headerMap]

    if missingHeaders:
        raise ValueError(
            "Cannot refresh summaries because Tasks is missing these columns: "
            + ", ".join(missingHeaders)
        )

    weekStartDates, monthStartDates = collectSummaryPeriodsFromTasks(wsTasks, headerMap)

    removeWorksheetIfExists(workbook, "Weekly Summary")
    removeWorksheetIfExists(workbook, "Monthly Summary")

    refreshWeeklySummarySheet(workbook, weekStartDates, headerMap)
    refreshMonthlySummarySheet(workbook, monthStartDates, headerMap)

    if "Tasks" in workbook.sheetnames:
        workbook.active = workbook.sheetnames.index("Tasks")

def createUpdatedWorkbookCopy(existingWorkbookPath: Path, outputPath: Path) -> Path:
    sourcePath = Path(existingWorkbookPath)
    outputPath = Path(outputPath)
    outputPath.parent.mkdir(parents=True, exist_ok=True)

    sourceResolvedPath = sourcePath.resolve()

    try:
        outputResolvedPath = outputPath.resolve()
    except FileNotFoundError:
        outputResolvedPath = outputPath.absolute()

    if sourceResolvedPath == outputResolvedPath:
        outputPath = sourcePath.with_name(f"{sourcePath.stem}_Updated{sourcePath.suffix}")

    shutil.copy2(sourcePath, outputPath)
    return outputPath


def parseExcelDateTimeValue(value) -> datetime:
    if value is None:
        return datetime.min

    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = str(value).strip()

    if not text:
        return datetime.min

    dateTimeFormats = [
        "%m/%d/%Y %H:%M",
        "%m/%d/%y %H:%M",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%y %I:%M %p",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
    ]

    for dateTimeFormat in dateTimeFormats:
        try:
            return datetime.strptime(text, dateTimeFormat)
        except ValueError:
            pass

    return datetime.min


def rebuildTaskRowFormulas(wsTasks, rowNum: int, headerMap: dict[str, int]) -> None:
    earningCol = headerMap["Earning"]
    actualMinutesCol = headerMap["Actual Minutes"]
    paidMinutesCol = headerMap["Paid Minutes"]
    taskTypeCol = headerMap["Type of Task"]
    projectCol = headerMap["Project"]

    actualMinutesColLetter = get_column_letter(actualMinutesCol)
    paidMinutesColLetter = get_column_letter(paidMinutesCol)
    taskTypeColLetter = get_column_letter(taskTypeCol)

    wsTasks.cell(row=rowNum, column=earningCol).value = (
        f'=IF(OR(${paidMinutesColLetter}{rowNum}="",\'Settings\'!$B$4=""),"",'
        f'${paidMinutesColLetter}{rowNum}/60*\'Settings\'!$B$4)'
    )

    timeCapCol = headerMap.get("Time Cap Minutes")

    if timeCapCol is not None:
        timeCapColLetter = get_column_letter(timeCapCol)

        timeCapLookupFormula = (
            f'VLOOKUP(${taskTypeColLetter}{rowNum},'
            f'\'Task Types\'!$A$3:$B$1000,2,FALSE)'
        )

        wsTasks.cell(row=rowNum, column=timeCapCol).value = (
            f'=IF(${taskTypeColLetter}{rowNum}="","",'
            f'IFERROR(IF({timeCapLookupFormula}=0,"",{timeCapLookupFormula}),""))'
        )

        wsTasks.cell(row=rowNum, column=paidMinutesCol).value = (
            f'=IF(OR(${actualMinutesColLetter}{rowNum}="",${timeCapColLetter}{rowNum}=""),"",'
            f'MIN(${actualMinutesColLetter}{rowNum},${timeCapColLetter}{rowNum}))'
        )
    else:
        timeCapLookupFormula = (
            f'VLOOKUP(${taskTypeColLetter}{rowNum},'
            f'\'Task Types\'!$A$3:$B$1000,2,FALSE)'
        )

        wsTasks.cell(row=rowNum, column=paidMinutesCol).value = (
            f'=IF(OR(${actualMinutesColLetter}{rowNum}="",${taskTypeColLetter}{rowNum}=""),"",'
            f'IFERROR(IF({timeCapLookupFormula}=0,"",'
            f'MIN(${actualMinutesColLetter}{rowNum},{timeCapLookupFormula})),""))'
        )

    wsTasks.cell(row=rowNum, column=projectCol).value = "='Settings'!$B$3"


def sortTasksSheetByDate(wsTasks, headerMap: dict[str, int], newestFirst: bool = True) -> int:
    taskIdCol = headerMap["Task ID"]
    dateCol = headerMap["Date"]
    maxCol = wsTasks.max_column

    formulaCols = {
        headerMap["Earning"],
        headerMap["Paid Minutes"],
        headerMap["Project"],
    }

    if "Time Cap Minutes" in headerMap:
        formulaCols.add(headerMap["Time Cap Minutes"])

    taskRows = []

    for rowNum in range(2, wsTasks.max_row + 1):
        taskIdValue = wsTasks.cell(row=rowNum, column=taskIdCol).value

        if taskIdValue is None or not str(taskIdValue).strip():
            continue

        rowValues = {}

        for colNum in range(1, maxCol + 1):
            if colNum in formulaCols:
                continue

            rowValues[colNum] = wsTasks.cell(row=rowNum, column=colNum).value

        taskRows.append(
            {
                "taskId": str(taskIdValue).strip(),
                "sortDateTime": parseExcelDateTimeValue(wsTasks.cell(row=rowNum, column=dateCol).value),
                "values": rowValues,
            }
        )

    taskRows.sort(
        key=lambda rowData: (rowData["sortDateTime"], rowData["taskId"]),
        reverse=newestFirst,
    )

    for rowNum in range(2, wsTasks.max_row + 1):
        for colNum in range(1, maxCol + 1):
            wsTasks.cell(row=rowNum, column=colNum).value = None

    for outputRowNum, rowData in enumerate(taskRows, start=2):
        for colNum, value in rowData["values"].items():
            wsTasks.cell(row=outputRowNum, column=colNum).value = value

        rebuildTaskRowFormulas(wsTasks, outputRowNum, headerMap)

    lastTaskRow = len(taskRows) + 1

    if wsTasks.max_row > lastTaskRow:
        wsTasks.delete_rows(lastTaskRow + 1, wsTasks.max_row - lastTaskRow)

    return len(taskRows)


def refreshTaskTypeDropdownRange(wsTasks, headerMap: dict[str, int]) -> None:
    if "Type of Task" not in headerMap:
        return

    taskTypeColLetter = get_column_letter(headerMap["Type of Task"])
    targetRange = f"{taskTypeColLetter}2:{taskTypeColLetter}{max(wsTasks.max_row + 500, 1000)}"

    if wsTasks.data_validations is not None:
        for dataValidation in wsTasks.data_validations.dataValidation:
            if dataValidation.type == "list" and str(dataValidation.formula1) == "=TaskTypeList":
                dataValidation.add(targetRange)
                return

    taskDropdown = DataValidation(
        type="list",
        formula1="=TaskTypeList",
        allow_blank=True,
        showDropDown=False,
    )
    taskDropdown.showErrorMessage = False
    taskDropdown.error = "Choose an approved task type from the dropdown. To add a new one, add it manually in Task Types columns A:B."
    taskDropdown.errorTitle = "Task type"
    taskDropdown.prompt = "Pick a task type from the dropdown. To add a new type, add it manually in Task Types columns A:B."
    taskDropdown.promptTitle = "Type of Task"

    wsTasks.add_data_validation(taskDropdown)
    taskDropdown.add(targetRange)

def getLastTaskTypeRow(wsTaskTypes) -> int:
    lastTaskTypeRow = 2

    for rowNum in range(3, wsTaskTypes.max_row + 1):
        taskTypeValue = wsTaskTypes.cell(row=rowNum, column=1).value

        if taskTypeValue is not None and str(taskTypeValue).strip():
            lastTaskTypeRow = rowNum

    return lastTaskTypeRow


def getExistingTaskTypeNames(wsTaskTypes) -> set[str]:
    existingTaskTypes = set()

    for rowNum in range(3, wsTaskTypes.max_row + 1):
        taskTypeValue = wsTaskTypes.cell(row=rowNum, column=1).value

        if taskTypeValue is not None and str(taskTypeValue).strip():
            existingTaskTypes.add(str(taskTypeValue).strip())

    return existingTaskTypes


def updateTaskTypesTableRange(workbook) -> None:
    if "Task Types" not in workbook.sheetnames:
        return

    wsTaskTypes = workbook["Task Types"]
    lastTaskTypeRow = getLastTaskTypeRow(wsTaskTypes)

    if lastTaskTypeRow < 2:
        return

    newRef = f"A2:C{lastTaskTypeRow}"

    for table in wsTaskTypes.tables.values():
        if table.displayName == "TaskTypesTable":
            table.ref = newRef
            return


def syncDefaultTaskTypesIntoWorkbook(workbook) -> None:
    """
    Keeps task types that the user manually added in the existing workbook,
    and also appends any new task types from defaultTimeCapsMin that are missing.
    Unknown caps can be None and will stay blank in Excel.
    """
    if "Task Types" not in workbook.sheetnames:
        return

    wsTaskTypes = workbook["Task Types"]
    existingTaskTypes = getExistingTaskTypeNames(wsTaskTypes)
    nextRow = getLastTaskTypeRow(wsTaskTypes) + 1

    for taskType, timeCapMinutes in defaultTimeCapsMin.items():
        taskTypeName = str(taskType).strip()

        if not taskTypeName or taskTypeName in existingTaskTypes:
            continue

        wsTaskTypes.cell(row=nextRow, column=1).value = taskTypeName
        wsTaskTypes.cell(row=nextRow, column=2).value = timeCapMinutes if timeCapMinutes is not None else None
        wsTaskTypes.cell(row=nextRow, column=3).value = "You can change this max paid time anytime."

        for colNum in range(1, 4):
            cell = wsTaskTypes.cell(row=nextRow, column=colNum)
            cell.alignment = Alignment(vertical="center")
            cell.border = thinBorder

        existingTaskTypes.add(taskTypeName)
        nextRow += 1

    updateTaskTypesTableRange(workbook)

def appendMissingTasksToExistingWorkbook(
    tasks: list[dict],
    projectName: str,
    hourlyPay: float,
    existingWorkbookPath: Path,
    outputPath: Path,
) -> Path:
    if not existingWorkbookPath.exists():
        raise FileNotFoundError(f"Existing Excel tracker was not found: {existingWorkbookPath}")

    updatedWorkbookPath = createUpdatedWorkbookCopy(
        existingWorkbookPath=existingWorkbookPath,
        outputPath=outputPath,
    )

    workbook = load_workbook(updatedWorkbookPath)

    syncDefaultTaskTypesIntoWorkbook(workbook)

    if "Tasks" not in workbook.sheetnames:
        raise ValueError(f"Existing workbook does not have a Tasks sheet: {existingWorkbookPath}")

    wsTasks = workbook["Tasks"]
    headerMap = getHeaderMap(wsTasks)

    requiredHeaders = [
        "Task ID",
        "Date",
        "Total Time",
        "Type of Task",
        "Earning",
        "Actual Minutes",
        "Paid Minutes",
        "Week Start",
        "Week End",
        "Project",
    ]

    missingHeaders = [header for header in requiredHeaders if header not in headerMap]

    if missingHeaders:
        raise ValueError(
            "Existing workbook is missing required Tasks columns: "
            + ", ".join(missingHeaders)
        )

    taskIdCol = headerMap["Task ID"]
    dateCol = headerMap["Date"]
    totalTimeCol = headerMap["Total Time"]
    taskTypeCol = headerMap["Type of Task"]
    actualMinutesCol = headerMap["Actual Minutes"]
    weekStartCol = headerMap["Week Start"]
    weekEndCol = headerMap["Week End"]

    earningCol = headerMap["Earning"]
    paidMinutesCol = headerMap["Paid Minutes"]
    projectCol = headerMap["Project"]
    timeCapCol = headerMap.get("Time Cap Minutes")

    existingTaskIds = getExistingTaskIds(wsTasks, taskIdCol)
    addedTaskCount = 0

    for task in tasks:
        taskId = str(task["taskId"]).strip()

        if not taskId or taskId in existingTaskIds:
            continue

        rowNum = wsTasks.max_row + 1

        wsTasks.cell(row=rowNum, column=taskIdCol).value = task["taskId"]
        wsTasks.cell(row=rowNum, column=dateCol).value = task["date"]
        wsTasks.cell(row=rowNum, column=totalTimeCol).value = task["totalTime"]
        wsTasks.cell(row=rowNum, column=taskTypeCol).value = ""
        wsTasks.cell(row=rowNum, column=actualMinutesCol).value = task["actualMinutes"]
        wsTasks.cell(row=rowNum, column=weekStartCol).value = task["weekStart"]
        wsTasks.cell(row=rowNum, column=weekEndCol).value = task["weekEnd"]

        rebuildTaskRowFormulas(wsTasks, rowNum, headerMap)

        wsTasks.cell(row=rowNum, column=dateCol).number_format = "m/d/yyyy"
        wsTasks.cell(row=rowNum, column=earningCol).number_format = '$#,##0.00'
        wsTasks.cell(row=rowNum, column=actualMinutesCol).number_format = "0.00"
        wsTasks.cell(row=rowNum, column=paidMinutesCol).number_format = "0.00"
        wsTasks.cell(row=rowNum, column=weekStartCol).number_format = "m/d/yyyy"
        wsTasks.cell(row=rowNum, column=weekEndCol).number_format = "m/d/yyyy"

        if timeCapCol is not None:
            wsTasks.cell(row=rowNum, column=timeCapCol).number_format = "0.00"

        wsTasks.cell(row=rowNum, column=projectCol).number_format = "General"

        existingTaskIds.add(taskId)
        addedTaskCount += 1

    sortedTaskCount = sortTasksSheetByDate(wsTasks, headerMap, newestFirst=True)
    lastTaskRow = sortedTaskCount + 1

    updateTasksTableRange(wsTasks, lastTaskRow=lastTaskRow)
    refreshTaskTypeDropdownRange(wsTasks, headerMap)
    refreshSummarySheets(workbook)

    if "Settings" in workbook.sheetnames:
        wsSettings = workbook["Settings"]

        if wsSettings.max_row >= 3:
            wsSettings["B3"].value = projectName

        if wsSettings.max_row >= 4:
            wsSettings["B4"].value = hourlyPay

    workbook.save(updatedWorkbookPath)

    print(f"Original Excel tracker was not edited: {existingWorkbookPath.resolve()}")
    print(f"Updated Excel tracker copy saved here: {updatedWorkbookPath.resolve()}")
    print(f"New tasks added into copied Excel tracker: {addedTaskCount}")
    print(f"Skipped duplicate task IDs: {len(tasks) - addedTaskCount}")
    print(f"Tasks sorted by date newest first: {sortedTaskCount}")

    return updatedWorkbookPath

def createHandshakeEarningsTracker(
    csvPath: str | Path | None = None,
    outputFolder: str = "Output",
    outputFileName: str = "HandshakeEarningTracker.xlsx",
    projectName: str | None = None,
    hourlyPay: float | None = None,
    existingWorkbookPath: str | Path | None = None,
    outputPath: str | Path | None = None,
) -> Path | None:
    """Creates a new tracker or updates a copied existing tracker with missing CSV task IDs only."""
    try:
        resolvedCsvPath = resolveCsvPath(str(csvPath) if csvPath else None, outputFolder)

        if not resolvedCsvPath.exists():
            raise FileNotFoundError(f"CSV file not found: {resolvedCsvPath}")

        if projectName is None or hourlyPay is None:
            csvProjectName, csvHourlyPay = readProjectDetailsFromCsv(resolvedCsvPath)
            projectName = projectName or csvProjectName
            hourlyPay = hourlyPay if hourlyPay is not None else csvHourlyPay

        tasks = readTasks(resolvedCsvPath)
        finalOutputPath = resolveOutputPath(outputPath, outputFolder, outputFileName)

        resolvedExistingWorkbookPath = resolveExistingWorkbookPath(
            existingWorkbookPath,
            outputFolder,
            outputFileName,
        )

        if resolvedExistingWorkbookPath is not None:
            return appendMissingTasksToExistingWorkbook(
                tasks=tasks,
                projectName=projectName,
                hourlyPay=float(hourlyPay),
                existingWorkbookPath=resolvedExistingWorkbookPath,
                outputPath=finalOutputPath,
            )

        createWorkbook(tasks, projectName, float(hourlyPay), finalOutputPath)

        print(f"Created Excel tracker: {finalOutputPath.resolve()}")
        print(f"Tasks imported into Excel: {len(tasks)}")

        return finalOutputPath

    except PermissionError as error:
        print("Excel creation failed: Permission denied. Close the Excel file if it is open, then try again.")
        print(f"Python error: {error}")
        return None

    except FileNotFoundError as error:
        print("Excel creation failed because a needed file could not be found.")
        print(f"Python error: {error}")
        return None

    except Exception as error:
        print("Excel creation failed because of an unexpected error.")
        print(f"Python error: {error}")
        return None

def main() -> None:
    print("haiExcelCreator.py is a helper module for creating or updating the Excel tracker.")
    print("Run main.py instead so the scraper, CSV creator, and Excel creator work together.")


if __name__ == "__main__":
    main()